#!/usr/bin/env python3
"""Adapter: MIT AI Risk Repository (taxonomy and risk database).

Unlike the other adapters this one does NOT produce candidate events. It
ingests reference data: two taxonomies and the labelled risk database that
uses them. The taxonomies are the shared vocabulary everything else gets
tagged against, which is what turns a flat pile of candidates into something
with navigable structure.

  Causal Taxonomy   Entity (AI/Human/Other) x Intent (Intentional/
                    Unintentional/Other) x Timing (Pre-/Post-deployment/Other)
  Domain Taxonomy   7 domains, 24 sub-domains
  Risk Database     2,500+ rows from 70+ source frameworks, each labelled
                    against both taxonomies

Licence: CC-BY-4.0, stated on airisk.mit.edu and inside the workbook itself
("This work is licensed under CC BY 4.0").

Requires openpyxl. Install with: pip install openpyxl

Usage:
    python scripts/adapters/mit_airr.py
    python scripts/adapters/mit_airr.py --dry-run

Writes an immutable dump under data/raw/mit_airr/dumps/<timestamp>/.
"""

import argparse
import json
import os
import sys

sys.path.insert(0, __file__.rsplit("mit_airr.py", 1)[0])

import _base  # noqa: E402

SOURCE_ID = "mit_airr"
ADAPTER_VERSION = "0.1.0"

SHEET_ID = "15LeHcpeuZC9txkvcaMoh3sUhkMvdMMry69xxXL46DT0"
XLSX_URL = ("https://docs.google.com/spreadsheets/d/%s/export?format=xlsx"
            % SHEET_ID)

LICENSE = {
    "spdx": "CC-BY-4.0",
    "url": "https://creativecommons.org/licenses/by/4.0/",
    "attribution": "MIT AI Risk Repository (MIT AI Risk Initiative)",
    "citation": (
        "Slattery, P., Saeri, A. K., Grundy, E. A. C., Graham, J., Noetel, M., "
        "Uuk, R., Dao, J., Pour, S., Casper, S., & Thompson, N. (2024). The AI "
        "Risk Repository: A Comprehensive Meta-Review, Database, and Taxonomy "
        "of Risks from Artificial Intelligence. arXiv:2408.12622"
    ),
    "source_terms_url": "https://airisk.mit.edu/",
    "verified_at": "2026-07-25",
    "verified_by": (
        "first-party site statement plus an in-artifact statement on the "
        "Domain Taxonomy sheet: 'This work is licensed under CC BY 4.0'"
    ),
}

SHEET_CAUSAL = "Causal Taxonomy of AI Risks v1"
SHEET_DOMAIN = "Domain Taxonomy of AI Risks v1"
SHEET_DATABASE = "AI Risk Database v4"
SHEET_RESOURCES = "Included resources"

DATABASE_HEADER_ROW = 3  # 1-indexed; rows 1-2 are titles and links


def clean(value):
    if value is None:
        return None
    text = _base.to_ascii(str(value)).strip()
    return " ".join(text.split()) or None


def fetch_workbook(session):
    response = _base.polite_get(session, XLSX_URL, timeout=120)
    return response.content, _base.sha256_bytes(response.content)


def parse_causal(ws):
    """Rows are (Category, Level, Description) with Category only on the first
    row of each block, so the current category carries downward."""
    out = []
    current = None
    for row in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in row]
        present = [c for c in cells if c]
        if len(present) < 2:
            continue
        if len(present) >= 3 and present[0] in ("Entity", "Intent", "Timing"):
            current = present[0]
            out.append({"category": current, "level": present[1],
                        "description": present[2]})
        elif current and len(present) >= 2 and present[0] != "Category":
            out.append({"category": current, "level": present[0],
                        "description": present[1]})
    return out


def parse_domain(ws):
    """Domain rows carry an id like '1.0' plus a name; sub-domain rows carry
    an id like '1.1'. Domain only appears on its first sub-domain row."""
    out = []
    current_id = None
    current_name = None
    for row in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in row]
        present = [c for c in cells if c]
        if len(present) < 2:
            continue
        head = present[0]
        if head and head.endswith(".0") and len(present) >= 4:
            current_id, current_name = head, present[1]
            out.append({
                "domain_id": current_id, "domain": current_name,
                "subdomain_id": present[2], "subdomain": present[3],
                "description": present[4] if len(present) > 4 else None,
            })
        elif head and "." in head and current_id and not head.endswith(".0"):
            out.append({
                "domain_id": current_id, "domain": current_name,
                "subdomain_id": head, "subdomain": present[1],
                "description": present[2] if len(present) > 2 else None,
            })
    return out


def parse_database(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = [clean(c) or "" for c in rows[DATABASE_HEADER_ROW - 1]]
    records = []
    for row in rows[DATABASE_HEADER_ROW:]:
        record = {}
        for index, name in enumerate(header):
            if name and index < len(row):
                record[name] = clean(row[index])
        if any(record.values()):
            records.append(record)
    return header, records


def parse_resources(ws):
    """The resources sheet has a preamble; the header is the row containing
    'Title'. Located by search rather than a fixed index, because a preamble
    edit upstream would silently shift a hardcoded offset."""
    rows = list(ws.iter_rows(values_only=True))
    header_index = None
    for index, row in enumerate(rows):
        cells = [clean(c) for c in row]
        if any(c and c.lower() == "title" for c in cells):
            header_index = index
            break
    if header_index is None:
        return [], []
    header = [clean(c) or "" for c in rows[header_index]]
    out = []
    for row in rows[header_index + 1:]:
        record = {}
        for index, name in enumerate(header):
            if name and index < len(row):
                record[name] = clean(row[index])
        if any(record.values()):
            out.append(record)
    return header, out


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. pip install openpyxl")
        return 1

    session = _base.get_session()
    blob, blob_sha = fetch_workbook(session)

    scratch = os.path.join(_base.REPO_ROOT, ".tmp_airr.xlsx")
    with open(scratch, "wb") as handle:
        handle.write(blob)
    try:
        workbook = openpyxl.load_workbook(scratch, data_only=True)
        causal = parse_causal(workbook[SHEET_CAUSAL])
        domain = parse_domain(workbook[SHEET_DOMAIN])
        db_header, database = parse_database(workbook[SHEET_DATABASE])
        res_header, resources = parse_resources(workbook[SHEET_RESOURCES])
        sheet_names = list(workbook.sheetnames)
    finally:
        os.remove(scratch)

    labelled = sum(1 for r in database if r.get("Domain") and r.get("Description"))
    domains = sorted(set(r["Domain"] for r in database if r.get("Domain")))

    print("causal taxonomy entries : %d" % len(causal))
    print("domain taxonomy entries : %d (%d domains)"
          % (len(domain), len(set(d["domain_id"] for d in domain))))
    print("risk database rows      : %d" % len(database))
    print("  with Domain + Description (usable as labelled data): %d" % labelled)
    print("included resources      : %d" % len(resources))
    print("distinct domains        : %d" % len(domains))

    if args.dry_run:
        print("\nfirst domain entry:", json.dumps(domain[0], sort_keys=True)[:260])
        print("first causal entry:", json.dumps(causal[0], sort_keys=True)[:200])
        print("\ndry run; nothing written")
        return 0

    dump_dir = os.path.join(_base.RAW_ROOT, SOURCE_ID, "dumps", _base.dump_stamp())
    os.makedirs(dump_dir, exist_ok=True)

    def write_json(name, payload):
        path = os.path.join(dump_dir, name)
        with open(path, "w", encoding="ascii", newline="\n") as handle:
            json.dump(payload, handle, ensure_ascii=True, indent=2, sort_keys=True)
            handle.write("\n")
        return name

    def write_jsonl(name, rows):
        path = os.path.join(dump_dir, name)
        with open(path, "w", encoding="ascii", newline="\n") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=True, sort_keys=True) + "\n")
        return name

    written = [
        write_json("taxonomy_causal.json", causal),
        write_json("taxonomy_domain.json", domain),
        write_jsonl("risk_database.jsonl", database),
        write_jsonl("included_resources.jsonl", resources),
    ]

    metadata = {
        "source_id": SOURCE_ID,
        "source_url": XLSX_URL,
        "source_name": "MIT AI Risk Repository",
        "extraction_method": "google_sheets_xlsx_export",
        "adapter_version": ADAPTER_VERSION,
        "license": LICENSE,
        "content_type": "reference_taxonomy_and_labelled_database",
        "content_note": (
            "NOT candidate events. This dump is reference data: the shared "
            "vocabulary other records get tagged against, plus the labelled "
            "database that vocabulary was derived from."
        ),
        "source_payload_sha256": blob_sha,
        "workbook_sheets": sheet_names,
        "extraction_statistics": {
            "causal_taxonomy_entries": len(causal),
            "domain_taxonomy_entries": len(domain),
            "risk_database_rows": len(database),
            "risk_database_labelled_rows": labelled,
            "included_resources": len(resources),
            "errors": 0,
        },
        "database_columns": db_header,
        "resources_columns": res_header,
        "tool_versions": {"python": sys.version.split()[0]},
        "notes": (
            "Upstream is a live Google Sheet updated in place, so the payload "
            "hash is the only fixed record of what a given dump saw. Sheet "
            "titles are version-stamped upstream (v1 taxonomies, v4 database); "
            "a bump there is a breaking change worth noticing."
        ),
    }
    write_json("_metadata.json", metadata)

    manifest_path = os.path.join(dump_dir, "MANIFEST.sha256")
    with open(manifest_path, "w", encoding="ascii", newline="\n") as handle:
        for name in sorted(written + ["_metadata.json"]):
            digest = _base.sha256_file(os.path.join(dump_dir, name))
            handle.write("%s  %s\n" % (digest, name))

    print("\nwrote dump: %s" % dump_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
